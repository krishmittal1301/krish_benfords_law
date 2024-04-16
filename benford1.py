import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np
import math
from collections import defaultdict
import matplotlib.pyplot as plt

# Read Excel file and extract column
df = pd.read_excel('GL for Expenses.xlsx')
df1 = df['Amt in local cur.']
x = (len(df1))
def get_first_digit(number):
    return int(str(number)[0])
    
def get_first_two_digit(number):
    return int(str(number)[:2])

def get_first_three_digit(number):
    return int(str(number)[:3])

def benfords1_law(data, i):
    counts = defaultdict(int)
    total = 0
    for value in data:
        first_two_digits = get_first_two_digit(value)
        counts[first_two_digits] += 1
        total += 1
    
    benfords = [math.log10(1 + 1 / digit) * total for digit in range(10 * i + 1, 10 * i + 10)]
    observed = [counts[digit] for digit in range(10 * i + 1, 10 * i + 10)]
    
    return benfords, observed

def benfords2_law(data,i,j):
    counts=defaultdict(int)
    total=0
    for value in data:
            first_three_digit=get_first_three_digit(value)
            counts[first_three_digit]+=1
            total+=1
    
    benfords=[math.log10(1 + 1/digit)* total for digit in range(100*i+10*j+1,100*i+10*j+10)]
    observed=[counts[digit] for digit in range(100*i+10*j+1,100*i+10*j+10)]

    return benfords, observed

def benfords_law(data):
    counts = defaultdict(int)
    total = 0
    for value in data:
        first_digit = get_first_digit(value)
        counts[first_digit] += 1
        total += 1
    
    benfords = [(math.log10(1 + 1 / digit) * total) for digit in range(1, 10)]
    observed = [counts[digit] for digit in range(1, 10)]
    return benfords, observed



# Calculate Benford's law deviations
benfords, observed = benfords_law(df1.values)
deviations = [((observed[i] - benfords[i]) / x) * 100 for i in range(9)]

print("Benford's Law Deviation:")
for digit, deviation in zip(range(1, 10), deviations):
    print(f"Digit {digit}: Deviation {deviation:.2f}%")

# Separate the positive and negative deviations
lstneg, lstpos = defaultdict(int), defaultdict(int)
for digit, deviation in zip(range(1, 10), deviations):
    if deviation < 0:
        lstneg[deviation] = digit
    elif deviation > 0:
        lstpos[deviation] = digit

# Sort the dictionaries
lstpos = sorted(lstpos.items(), reverse=True)
lstneg = sorted(lstneg.items())

# Extract the digits from the sorted lists
lst1 = [value for key, value in lstpos]

# Plot the subplots
maxdev=defaultdict(int)
for key ,value in  lstpos:
    benfords,observed=benfords1_law(df1.values,value)
    deviations = [((observed[i] - benfords[i]) / x) * 100 for i in range(9)]
    listx=[i for i in range (10*value +1 ,10*value+10)]
    plt.plot(listx,deviations)
    # plt.plot(listx,observed)
    plt.title("benford distribution of the values whose numbers are more than required")
    for i in range(len(deviations)):
        if (deviations[i]>0):
            maxdev[deviations[i]]=10*value+i

    
plt.show()
maxdev=sorted(maxdev.items())
for key,value in maxdev:
    print(f"deviation {key} : {value}")

threedev=defaultdict(int)
# for checking three digit numbers
# for key,value in maxdev:
#     b=str(value)
#     benfords,observed=benfords2_law(df.values,int(b[0]),int(b[1]))
#     deviations= [(observed[i] - benfords[i]) / benfords[i] * 100 for i in range(9)]
#     listx=[i for i in range(100*int(b[0])+10*int(b[1])+1 , 100*int(b[0])+10*int(b[1])+10)]
#     plt.plot(listx,deviations)
#     for i in range(len(deviations)):
#         if (deviations[i]>50):
#             threedev[deviations[i]]=100*int(b[0])+10*int(b[1])+i


column_to_check = 'Amt in local cur.'  
color = 'FFFF00'  

# # Load the existing workbook
wb = load_workbook('GL for Expenses.xlsx')
ws = wb.active

# Get the column index of the column to check
column_index = df.columns.get_loc(column_to_check) + 1

# Iterate through the column and apply formatting
for key,val in maxdev:
    for r_idx, value in enumerate(df[column_to_check], 1):
        cell = ws.cell(row=r_idx, column=column_index)
        cell.value = value
        
        # Check if the cell starts with "12" and apply color
        if str(value).startswith(str(val)):
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# Save the modified workbook
modified_file_path = 'modified_file.xlsx'  # Replace with your desired output file path
wb.save(modified_file_path)

